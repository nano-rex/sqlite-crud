#!/bin/python3

import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
import pandas as pd
import openpyxl
import shutil

# Root window
root = tk.Tk()
root.title('Display a Text File')
root.resizable(False, False)
root.geometry('550x250')

# Text editor
text = tk.Text(root, height=12)
text.grid(column=0, row=0, sticky='nsew')

def open_text_file():
    # file type
    filetypes = (
        ('Excel files', '*.xlsx'),
        ('All files', '*.*')
    )
    # show the open file dialog
    filepath = fd.askopenfilename(filetypes=filetypes)
    
    if filepath:
        # Create the new file path with "INV_" as the prefix
        def add_prefix_to_last_word(text, prefix):
           words = text.split("/") # Split the text by forward slashes
           last_word = words[-1] # Get the last word after the last forward slash
           new_last_word = prefix + last_word # Add the prefix to the last word
           words[-1] = new_last_word # Replace the last word with the updated one 
           modified_text = "/".join(words) # Join the words back together with forward slashes
           return modified_text
        text = filepath
        prefix = "INV_"
        new_filepath = add_prefix_to_last_word(filepath, prefix)
        shutil.copyfile(filepath, new_filepath)
        
        # Update the text widget with the new file path
        text.delete('1.0', tk.END)
        text.insert(tk.END, new_filepath)

def xlsx(): 
     shutil.copyfile('/home/user/Templates/INV_GCH.xlsx', new_filepath) # Copy Invoice template as destination xlsx

     df = pd.read_excel(filepath, sheet_name='Sheet1', header=None)
     # Read(copy) Work Order
     inv_no = df.loc[10,7]
     date_ = df.loc[12,7]
     po_no = df.loc[14,7]
     #info0 = df.loc[9,1]
     #info1 = df.loc[10,1]
     #info2 = df.loc[11,1]
     #info3 = df.loc[12,1]
     #info4 = df.loc[13,1]
     #info5 = df.loc[14,1]
     #info6 = df.loc[15,1]
     #info7 = df.loc[16,1]
     #info8 = df.loc[17,1]
     desc0 = df.loc[21,1]
     desc1 = df.loc[22,1]
     desc2 = df.loc[23,1]
     desc3 = df.loc[24,1]
     desc4 = df.loc[25,1]
     desc5 = df.loc[26,1]
     desc6 = df.loc[27,1]
     desc7 = df.loc[28,1]
     desc8 = df.loc[29,1]
     desc9 = df.loc[30,1]
     desc10 = df.loc[31,1]
     desc11 = df.loc[32,1]
     desc12 = df.loc[33,1]
     desc13 = df.loc[34,1]
     desc14 = df.loc[35,1]
     desc15 = df.loc[36,1]
     desc16 = df.loc[37,1]
     desc17 = df.loc[38,1]
     desc18 = df.loc[39,1]
     desc19 = df.loc[40,1]

     srcfile = openpyxl.load_workbook(new_filepath, read_only=False, keep_vba=False)
     sheetname = srcfile['Sheet1']
     # Write(paste) to Invoice
     sheetname['H11'] = str(inv_no)
     sheetname['H13'] = str(date_)
     sheetname['H15'] = str(po_no)
     #sheetname['B11'] = str(info0)
     #sheetname['B12'] = str(info1)
     #sheetname['B13'] = str(info2)
     #sheetname['B14'] = str(info3)
     #sheetname['B15'] = str(info4)
     #sheetname['B16'] = str(info5)
     #sheetname['B17'] = str(info6)
     #sheetname['B18'] = str(info7)
     sheetname['B23'] = str(desc0)
     sheetname['B24'] = str(desc1)
     sheetname['B25'] = str(desc2)
     sheetname['B26'] = str(desc3)
     sheetname['B27'] = str(desc4)
     sheetname['B28'] = str(desc5)
     sheetname['B29'] = str(desc6)
     sheetname['B30'] = str(desc7)
     sheetname['B31'] = str(desc8)
     sheetname['B32'] = str(desc9)
     sheetname['B33'] = str(desc10)
     sheetname['B34'] = str(desc11)
     sheetname['B35'] = str(desc12)
     sheetname['B36'] = str(desc13) # Site address
     sheetname['B37'] = str(desc14)
     sheetname['B38'] = str(desc15)
     sheetname['B39'] = str(desc16)
     sheetname['B40'] = str(desc17)
     sheetname['B41'] = str(desc18)
     sheetname['B42'] = str(desc19)
     srcfile.save(new_filepath)

     # Find and replace/remove "nan"
     sheet1 = srcfile.active
     for col in sheet1.iter_cols(min_col=0, min_row=0, max_col=9, max_row=56):
       for cell in col:
         if cell.value == 'nan':
           cell.value = ''
     srcfile.save(new_filepath)

     # Insert Logo image to Invoice
     wb = openpyxl.load_workbook(new_filepath)
     ws = wb['Sheet1']
     img = openpyxl.drawing.image.Image('/home/user/Templates/demax_logo.png')
     img.anchor = 'A3' # Or whatever cell location you want to use.
     ws.add_image(img)
     img1 = openpyxl.drawing.image.Image('/home/user/Templates/stella_signature_small.png')
     img1.anchor = 'G47'
     ws.add_image(img1)
     img2 = openpyxl.drawing.image.Image('/home/user/Templates/demax_stamp_small.png')
     img2.anchor = 'H46'
     ws.add_image(img2)
     wb.save(new_filepath)

# open file button
open_button = ttk.Button(
    root,
    text='Open a File',
    command=open_text_file
)

open_button.grid(column=0, row=1, sticky='w', padx=10, pady=10)

# run the application
root.mainloop()
