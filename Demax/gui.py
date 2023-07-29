#!/bin/python3

import pandas as pd
import openpyxl
import sys
import shutil
import tkinter
from tkinter import filedialog

filepath1 = ""

def add_prefix_to_last_word(text, prefix):
    words = text.split("/")
    last_word = words[-1]
    new_last_word = prefix + last_word
    words[-1] = new_last_word
    modified_text = "/".join(words)
    return modified_text

def open_folder():
    global filepath1
    filepath1 = filedialog.askopenfilename(initialdir="/home/user/Desktop", filetypes=[("Excel Files", "*.xlsx")])

def gch():
    global filepath1
    filepath2 = add_prefix_to_last_word(filepath1, "INV_")
    shutil.copyfile('/home/user/Templates/INV_GCH.xlsx', filepath2)
    df = pd.read_excel(filepath1, sheet_name='Sheet1', header=None)
    inv_no = df.loc[10,7]
    date_ = df.loc[12,7]
    po_no = df.loc[14,7]
    desc0 = df.loc[21,1]
    desc1 = df.loc[22,1]
    desc2 = df.loc[23,1]
    desc3 = df.loc[24,1]
    desc4 = df.loc[25,1]
    desc5 = df.loc[26,1]
    desc6 = df.loc[27,1]
    desc7 = df.loc[28,1]
    desc8 = df.loc[29,1]
    desc9 = df.loc[30,1]
    desc10 = df.loc[31,1]
    desc11 = df.loc[32,1]
    desc12 = df.loc[33,1]
    desc13 = df.loc[34,1]
    desc14 = df.loc[35,1]
    desc15 = df.loc[36,1]
    desc16 = df.loc[37,1]
    desc17 = df.loc[38,1]
    desc18 = df.loc[39,1]
    desc19 = df.loc[40,1]
    srcfile = openpyxl.load_workbook(filepath2, read_only=False, keep_vba=False)
    sheetname = srcfile['Sheet1']
    sheetname['H11'] = str(inv_no)
    sheetname['H13'] = str(date_)
    sheetname['H15'] = str(po_no)
    sheetname['B23'] = str(desc0)
    sheetname['B24'] = str(desc1)
    sheetname['B25'] = str(desc2)
    sheetname['B26'] = str(desc3)
    sheetname['B27'] = str(desc4)
    sheetname['B28'] = str(desc5)
    sheetname['B29'] = str(desc6)
    sheetname['B30'] = str(desc7)
    sheetname['B31'] = str(desc8)
    sheetname['B32'] = str(desc9)
    sheetname['B33'] = str(desc10)
    sheetname['B34'] = str(desc11)
    sheetname['B35'] = str(desc12)
    sheetname['B36'] = str(desc13)
    sheetname['B37'] = str(desc14)
    sheetname['B38'] = str(desc15)
    sheetname['B39'] = str(desc16)
    sheetname['B40'] = str(desc17)
    sheetname['B41'] = str(desc18)
    srcfile.save(filepath2)
    sheet1 = srcfile.active
    for col in sheet1.iter_cols(min_col=0, min_row=0, max_col=9, max_row=56):
        for cell in col:
            if cell.value == 'nan':
                cell.value = ''
    srcfile.save(filepath2)
    wb = openpyxl.load_workbook(filepath2)
    ws = wb['Sheet1']
    img = openpyxl.drawing.image.Image('/home/user/Templates/demax_logo.png')
    img.anchor = 'A3'
    ws.add_image(img)
    img1 = openpyxl.drawing.image.Image('/home/user/Templates/stella_signature_small.png')
    img1.anchor = 'G47'
    ws.add_image(img1)
    img2 = openpyxl.drawing.image.Image('/home/user/Templates/demax_stamp_small.png')
    img2.anchor = 'H46'
    ws.add_image(img2)
    wb.save(filepath2)
    success_message = "GCH invoice generated successfully!"
    output_path = "Output path: " + filepath2
    success_label.config(text=success_message)
    output_label.config(text=output_path)

def d_o():
    global filepath1
    filepath2 = add_prefix_to_last_word(filepath1, "INV_")
    shutil.copyfile('/home/user/Templates/INV_.xlsx', filepath2)
    df = pd.read_excel(filepath1, sheet_name='Sheet1', header=None)
    inv_no = df.loc[9,7]
    date_ = df.loc[11,7]
    po_no = df.loc[13,7]
    info0 = df.loc[9,1]
    info1 = df.loc[10,1]
    info2 = df.loc[11,1]
    info3 = df.loc[12,1]
    info4 = df.loc[13,1]
    info5 = df.loc[14,1]
    info6 = df.loc[15,1]
    info7 = df.loc[16,1]
    info8 = df.loc[17,1]
    desc0 = df.loc[22,1]
    desc1 = df.loc[23,1]
    desc2 = df.loc[24,1]
    desc3 = df.loc[25,1]
    desc4 = df.loc[26,1]
    desc5 = df.loc[27,1]
    desc6 = df.loc[28,1]
    desc7 = df.loc[29,1]
    desc8 = df.loc[30,1]
    desc9 = df.loc[31,1]
    desc10 = df.loc[32,1]
    desc11 = df.loc[33,1]
    desc12 = df.loc[34,1]
    desc13 = df.loc[35,1]
    desc14 = df.loc[36,1]
    desc15 = df.loc[37,1]
    desc16 = df.loc[38,1]
    desc17 = df.loc[39,1]
    desc18 = df.loc[40,1]
    srcfile = openpyxl.load_workbook(filepath2, read_only=False, keep_vba=False)
    sheetname = srcfile['Sheet1']
    sheetname['H11'] = str(inv_no)
    sheetname['H13'] = str(date_)
    sheetname['H15'] = str(po_no)
    sheetname['B11'] = str(info0)
    sheetname['B12'] = str(info1)
    sheetname['B13'] = str(info2)
    sheetname['B14'] = str(info3)
    sheetname['B15'] = str(info4)
    sheetname['B16'] = str(info5)
    sheetname['B17'] = str(info6)
    sheetname['B18'] = str(info7)
    sheetname['B23'] = str(desc8)
    sheetname['B24'] = str(desc1)
    sheetname['B25'] = str(desc2)
    sheetname['B26'] = str(desc3)
    sheetname['B27'] = str(desc4)
    sheetname['B28'] = str(desc5)
    sheetname['B29'] = str(desc6)
    sheetname['B30'] = str(desc7)
    sheetname['B31'] = str(desc8)
    sheetname['B32'] = str(desc9)
    sheetname['B33'] = str(desc10)
    sheetname['B34'] = str(desc11)
    sheetname['B35'] = str(desc12)
    sheetname['B36'] = str(desc13)
    sheetname['B37'] = str(desc14)
    sheetname['B38'] = str(desc15)
    sheetname['B39'] = str(desc16)
    sheetname['B40'] = str(desc17)
    sheetname['B41'] = str(desc18)
    srcfile.save(filepath2)
    sheet1 = srcfile.active
    for col in sheet1.iter_cols(min_col=0, min_row=0, max_col=9, max_row=56):
        for cell in col:
            if cell.value == 'nan':
                cell.value = ''
    srcfile.save(filepath2)
    wb = openpyxl.load_workbook(filepath2)
    ws = wb['Sheet1']
    img = openpyxl.drawing.image.Image('/home/user/Templates/demax_logo.png')
    img.anchor = 'A3'
    ws.add_image(img)
    img1 = openpyxl.drawing.image.Image('/home/user/Templates/stella_signature_small.png')
    img1.anchor = 'G47'
    ws.add_image(img1)
    img2 = openpyxl.drawing.image.Image('/home/user/Templates/demax_stamp_small.png')
    img2.anchor = 'H46'
    ws.add_image(img2)
    wb.save(filepath2)
    success_message = "D.O invoice generated successfully!"
    output_path = "Output path: " + filepath2
    success_label.config(text=success_message)
    output_label.config(text=output_path)

def dm():
    global filepath1
    filepath2 = add_prefix_to_last_word(filepath1, "INV_")
    shutil.copyfile('/home/user/Templates/INV_.xlsx', filepath2)
    df = pd.read_excel(filepath1, sheet_name='Sheet1', header=None)
    no_ = df.loc[10,8]
    date_ = df.loc[12,8]
    work_no = df.loc[14,8]
    job_no = df.loc[16,8]
    desc0 = df.loc[25,1]
    desc1 = df.loc[26,1]
    desc2 = df.loc[27,1]
    desc3 = df.loc[28,1]
    desc4 = df.loc[29,1]
    desc5 = df.loc[30,1]
    desc6 = df.loc[31,1]
    desc7 = df.loc[32,1]
    desc8 = df.loc[33,1]
    desc9 = df.loc[34,1]
    desc10 = df.loc[35,1]
    desc11 = df.loc[36,1]
    desc12 = df.loc[37,1]
    desc13 = df.loc[38,1]
    desc14 = df.loc[39,1]
    desc15 = df.loc[40,1]
    desc16 = df.loc[41,1]
    desc17 = df.loc[42,1]
    desc18 = df.loc[43,1]
    desc19 = df.loc[44,1]
    srcfile = openpyxl.load_workbook(filepath2, read_only=False, keep_vba=True)
    sheetname = srcfile['Sheet1']
    sheetname['I12'] = str(date_)
    sheetname['B27'] = str(date_)[2:]
    sheetname['I14'] = str(no_)
    sheetname['C27'] = str(no_)[2:]
    sheetname['D27'] = str(job_no)[2:]
    sheetname['G27'] = str(work_no)[2:]
    sheetname['B30'] = str(desc0)
    sheetname['B31'] = str(desc1)
    sheetname['B32'] = str(desc2)
    sheetname['B33'] = str(desc3)
    sheetname['B34'] = str(desc4)
    sheetname['B35'] = str(desc5)
    sheetname['B36'] = str(desc6)
    sheetname['B37'] = str(desc7)
    sheetname['B38'] = str(desc8)
    sheetname['B39'] = str(desc9)
    sheetname['B40'] = str(desc10)
    sheetname['B41'] = str(desc11)
    sheetname['B42'] = str(desc12)
    sheetname['B43'] = str(desc13)
    sheetname['B44'] = str(desc14)
    sheetname['B45'] = str(desc15)
    sheetname['B46'] = str(desc16)
    sheetname['B47'] = str(desc17)
    sheetname['B48'] = str(desc18)
    sheetname['B49'] = str(desc19)
    srcfile.save(filepath2)
    sheet1 = srcfile.active
    for col in sheet1.iter_cols(min_col=0, min_row=0, max_col=9, max_row=56):
        for cell in col:
            if cell.value == 'nan':
                cell.value = ''
    srcfile.save(filepath2)
    wb = openpyxl.load_workbook(filepath2)
    ws = wb['Sheet1']
    img = openpyxl.drawing.image.Image('/home/user/Templates/demax_logo.png')
    img.anchor = 'A3'
    ws.add_image(img)
    img1 = openpyxl.drawing.image.Image('/home/user/Templates/stella_signature_small.png')
    img1.anchor = 'G47'
    ws.add_image(img1)
    img2 = openpyxl.drawing.image.Image('/home/user/Templates/demax_stamp_small.png')
    img2.anchor = 'H46'
    ws.add_image(img2)
    wb.save(filepath2)
    success_message = "DM invoice generated successfully!"
    output_path = "Output path: " + filepath2
    success_label.config(text=success_message)
    output_label.config(text=output_path)

def open_gui():
    global filepath1
    root = tkinter.Tk()
    root.title("Demax Invoice Generator")

    # Add text above the open folder button
    text_above_button = tkinter.Label(root, text="Please select an Excel file:")
    text_above_button.pack()

    def open_folder_button_click():
        open_folder()

    open_folder_button = tkinter.Button(root, text="Open Folder", command=open_folder_button_click)
    open_folder_button.pack()

    # Add text below the open folder button
    text_below_button = tkinter.Label(root, text="Click one of the buttons below to generate the invoice:")
    text_below_button.pack()

    # Create a frame to hold the buttons
    button_frame = tkinter.Frame(root)
    button_frame.pack()

    def gch_button_click():
        gch()

    def d_o_button_click():
        d_o()

    def dm_button_click():
        dm()

    gch_button = tkinter.Button(button_frame, text="GCH", command=gch_button_click)
    gch_button.pack(side=tkinter.LEFT)

    d_o_button = tkinter.Button(button_frame, text="D.O", command=d_o_button_click)
    d_o_button.pack(side=tkinter.LEFT)

    dm_button = tkinter.Button(button_frame, text="DM", command=dm_button_click)
    dm_button.pack(side=tkinter.LEFT)

    # Add success message label
    success_label = tkinter.Label(root, text="")
    success_label.pack()

    # Add output path label
    output_label = tkinter.Label(root, text="")
    output_label.pack()

    root.mainloop()

open_gui()
